from selenium import webdriver
import os
import datetime, calendar
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

shops = {'Lotteria - TTTM Menas Mall Saigon Airport' : 0, 'BÚN CHẢ QUỲNH HÀ NỘI - Menas Mall' : 1, 'SUNSHINE DETOX - Menas Mall' : 2, 'Trà sữa Cing Hu Tang - Menas Mall' : 3, 
         'Bà Hội' : 7, 'Chè Má Việt' : 8, 'L\'amuse Coffee' : 9, 'Mena Gourmet Market - Menas Mall Saigon Airport' : 12, 'Steakhouse The Fan Kim Long Villas' : 13}

# shops = {'Lotteria - TTTM Menas Mall Saigon Airport' : 0, 'BÚN CHẢ QUỲNH HÀ NỘI - Menas Mall' : 1, 'SUNSHINE DETOX - Menas Mall' : 2, 'Trà sữa Cing Hu Tang - Menas Mall' : 3,
#         'Sky Shop' : 4, 'Mena Goumert Makert' : 5, 'The Fan Steak House' : 6, 'Bà Hội' : 7, 'Chè Má Việt' : 8, 'L\'amuse Coffee' : 9, 'Yum Goumert Station' : 10,
#         'Mena World' : 11, 'Mena Gourmet Market - Menas Mall Saigon Airport' : 12, 'Steakhouse The Fan Kim Long Villas' : 13}

shops_code = {0 : 'Lotteria', 2 : 'SUNSHINE DETOX', 3 : 'CING HU TANG', 13 : 'SteakHouse The Fan', 7 : 'BÀ HỘI', 8 : 'CHÈ MÁ VIỆT', 9 : 'L\'amuse coffee'}

def get_time():
    today = datetime.date.today()-datetime.timedelta(1)
    return today, today.day, today.month, today.year

def get_month_leght(year, month):
    return calendar.monthrange(year, month)[1]    

def create_header(num_days):
    header = ['', '', 'TOTAL']
    for name in shops_code.values():
        for _ in range(3):
            header.append(name)
    
    header2 = ['No', 'Date', f'=SUM(C3:C{num_days + 2})']
    for _ in range(len(shops_code)):
        header2.append('Orders')
        header2.append('Net Revenue')
        header2.append('Total Sales')
    return header, header2

def read_csv(day, today):
    df = pd.read_csv('./Today_Sale.csv')
    data = []
    data = [int(day), today.strftime('%d/%m/%Y')]
    total = 0
    for index, row in df.iterrows():
        shop_name = row['Shop Name']
        shop_index = shops[shop_name]
        if shop_index in shops_code.keys():
            data.append(row['Number of Order'])
            data.append(int(row['Net'].replace(' ₫', '').replace('.', '')))
            data.append(int(row['Gross'].replace(' ₫', '').replace('.', '')))
            total += int(row['Gross'].replace(' ₫', '').replace('.', ''))
    data.insert(2, total)
    return data

def create_df():
    columns = ['Shop Name', 'Gross', 'Net', 'Number of Order']
    df = pd.DataFrame(columns = columns)
    return df

def save_df(df):
    if os.path.exists("Today_Sale.csv"):
        os.remove("Today_Sale.csv")
    df.to_csv('Today_Sale.csv', index=False)

def wait():
    time.sleep(1)

def create_excel(header, header2, year, month, num_days, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Grabfood'

    ws.append(header)
    for i in range(4, 24, 3):
        char = get_column_letter(i)
        last_char = get_column_letter(i+2)
        ws.merge_cells(f'{char}1:{last_char}1')

    ws.append(header2)

    for i in range(1, num_days + 1):
        ws.append([i, datetime.date(year, month, i).strftime('%d/%m/%Y')])

    for i in range(4, 25):
        char = get_column_letter(i)
        ws[char + str(num_days + 3)] = f'=SUM({char}3:{char}{num_days + 2})'

    formula = '=SUM('
    for i in range(len(shops_code)):
        formula = formula + f'{get_column_letter(4 + 3 * i)}{num_days + 3} + '
    formula = formula + '0 )'
    ws.append(['', '', 'Tổng Order',formula])

    wb.save(filename)

def load_excel(filename):
    wb = load_workbook(filename)
    ws = wb['Grabfood']

    return wb, ws

def insert_data(wb, ws, data):
    for i in range(2, int(len(shops_code)*3 + 3)):
        char = get_column_letter(i+1)
        if i != 2:
            ws[char + str(data[0] + 2)] = data[i]
        else:
            formula = '=SUM('
            for j in range(len(shops_code)):
                formula = formula + f'{get_column_letter(i + 3 + 3 * j)}{data[0] + 2} + '
            formula = formula + '0 )'
            ws[char + str(data[0] + 2)] = formula
    wb.save(filename)

def scrap_data(day):
    try:
        df = create_df()
        service = Service(executable_path = 'chromedriver.exe')
        driver = webdriver.Chrome(service = service)
        driver.get('https://merchant.grab.com/portal')
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CLASS_NAME, 'accept-btn-container')))
        driver.find_element(By.CLASS_NAME, 'accept-btn-container').click()
        driver.find_element(By.ID, 'username').send_keys('tai.ecommerce')
        driver.find_element(By.ID, 'password').send_keys('212211wwW')
        driver.find_element(By.CSS_SELECTOR, "button.dui-btn.css-eg1a4l.dui-btn-round.dui-btn-primary.dui-btn-lg.dui-btn-block").click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div/div[2]/div/div[2]/div/div[2]/button[2]")))
        wait()
        driver.find_element(By.XPATH, "/html/body/div[3]/div/div[2]/div/div[2]/div/div[2]/button[2]").click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/div/div[2]/div/div[2]/div/div/div/div[3]/button[1]")))
        wait()
        driver.find_element(By.XPATH, "/html/body/div[5]/div/div[2]/div/div[2]/div/div/div/div[3]/button[1]").click()
        wait()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div/div[2]/div/div/div[1]/div[4]/div[2]/div[1]/div[1]/div[1]/div[1]/div')))
        driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div[2]/div/div/div[1]/div[4]/div[2]/div[1]/div[1]/div[1]/div[1]/div').click()
        wait()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="insight-filter-container"]/div[1]')))
        driver.find_element(By.XPATH, '//*[@id="insight-filter-container"]/div[1]').click()
        wait()
        dropdown = driver.find_element(By.XPATH, '//div[@class="rc-virtual-list-holder-inner"] /div[@title = "mex-insightsv2-003-011-dropdown"]')
        ActionChains(driver).move_to_element(dropdown).click(dropdown).perform()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, f"//td[@role = 'gridcell']/div[text() = '{day}']")))
        wait()
        driver.find_element(By.XPATH, f"//td[@role = 'gridcell']/div[text() = '{day}']").click()
        wait()
        driver.find_element(By.XPATH, f"//td[@role = 'gridcell']/div[text() = '{day}']").click()
        wait()
        driver.find_element(By.XPATH, "//button/span[text() = 'Apply']").click()
        wait()
        shop_select = driver.find_element(By.XPATH, '//*[@id="insight-filter-container"]/div[3]/div')
        shop_select.click()
        wait()
        for shop in shops.keys():
            shop_sale = driver.find_element(By.XPATH, f'//div[@role = "option"][@title = "{shop}"]')
            ActionChains(driver).move_to_element(shop_sale).click(shop_sale).perform()
            driver.find_element(By.XPATH, f'//*[@id="root"]/div/div/div[2]/div/div/div[1]/div[2]/div[3]/div/div/div[2]/div[2]/div[1]/div[1]').click()
            time.sleep(2)
            gross = driver.find_element(By.XPATH, f'//*[@id="root"]/div/div/div[2]/div/div/div[1]/div[2]/div[3]/div/div/div[2]/div[2]/div[1]/div[1]/div[2]').text
            net = driver.find_element(By.XPATH, f'//*[@id="root"]/div/div/div[2]/div/div/div[1]/div[2]/div[3]/div/div/div[2]/div[2]/div[1]/div[2]/div[2]').text
            order_num = driver.find_element(By.XPATH, f'//*[@id="root"]/div/div/div[2]/div/div/div[1]/div[2]/div[3]/div/div/div[2]/div[2]/div[1]/div[3]/div[2]').text
            data  = {'Shop Name' : shop, 'Gross' : gross, 'Net' : net, 'Number of Order' : order_num}
            df.loc[len(df.index)] = data
            wait()
            shop_select.click()
            wait()
            try:
                ActionChains(driver).move_to_element(shop_sale).click(shop_sale).perform()
            except:
                wait()
                shop_select.click()
            wait()
        return True, df
    except:
        return False, df
    

if __name__ == "__main__":
    today, day, month, year = get_time()
    flag = False
    
    while flag != True:
        flag, df = scrap_data(day)
        
    save_df(df)

    data = read_csv(day, today)

    filename = f'Total_Sales_{month}.xlsx'

    if not os.path.exists(filename):
        num_days = get_month_leght(year, month)
        header, header2 = create_header(num_days)
        create_excel(header, header2, year, month, num_days, filename)

    wb, ws = load_excel(filename)

    insert_data(wb, ws, data)

    
